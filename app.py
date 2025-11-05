import os
import shutil
import tempfile
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file, session
from werkzeug.utils import secure_filename
import pickle
import asyncio
import io

# Import processing modules
from textract_processor import extract_structured_data_from_pdf_bytes
from structured_llm_processor import (
    process_structured_data_with_llm_async, 
    set_custom_context_prompt, 
    clear_custom_context_prompt,
    get_optimization_stats
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Use environment variable for secret key in production
app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-change-in-production')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    """Health check endpoint for Railway"""
    return jsonify({'status': 'healthy', 'service': 'PDF Extractor'})

@app.route('/extract', methods=['POST'])
def extract():
    """Extract structured data from uploaded PDF"""
    try:
        print("Extract endpoint called")
        print(f"Request files: {list(request.files.keys())}")
        
        if 'pdf' not in request.files:
            return jsonify({'error': 'No PDF file provided'}), 400
        
        file = request.files['pdf']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        print(f"File received: {file.filename}, size: {len(file.read())}")
        file.seek(0)  # Reset file pointer
        
        # Read PDF bytes
        pdf_bytes = file.read()
        print(f"PDF bytes read: {len(pdf_bytes)} bytes")
        
        # Extract structured data
        print("Calling extract_structured_data_from_pdf_bytes...")
        structured_data = extract_structured_data_from_pdf_bytes(pdf_bytes)
        print(f"Structured data extracted: {len(structured_data)} items")
        
        # Store file data in session for prompt testing
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pkl') as temp_file:
            pickle.dump(structured_data, temp_file)
            session['file_data_path'] = temp_file.name
            print(f"Stored file data at: {temp_file.name}")
        
        return jsonify({
            'success': True,
            'data': structured_data,
            'message': f'Successfully extracted data from {file.filename}'
        })
        
    except Exception as e:
        print(f"Error in extract endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process_stream', methods=['POST'])
def process_stream():
    """Process structured data with streaming response"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        def generate():
            try:
                # Process with LLM asynchronously
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                
                try:
                    result = loop.run_until_complete(
                        process_structured_data_with_llm_async(data)
                    )
                finally:
                    loop.close()
                
                # Send header
                yield f"data: {{'type': 'header', 'content': 'Field,Value,Context'}}\n\n"
                
                # Send data rows
                if 'enhanced_data_with_comprehensive_context' in result:
                    for i, entity in enumerate(result['enhanced_data_with_comprehensive_context']):
                        field = str(entity.get('field', 'Unknown')).replace('"', '""')
                        value = str(entity.get('value', '')).replace('"', '""')
                        context = str(entity.get('context', entity.get('Context', ''))).replace('"', '""')
                        
                        csv_row = f'row{i+1}: "{field}","{value}","{context}"'
                        yield f"data: {{'type': 'row', 'content': '{csv_row}'}}\n\n"
                
                # Send completion
                total_rows = len(result.get('enhanced_data_with_comprehensive_context', []))
                cost_summary = result.get('cost_summary', {})
                yield f"data: {{'type': 'complete', 'total_rows': {total_rows}, 'cost_summary': {cost_summary}}}\n\n"
                
            except Exception as e:
                yield f"data: {{'status': 'error', 'error': '{str(e)}'}}\n\n"
        
        return app.response_class(generate(), mimetype='text/plain')
        
    except Exception as e:
        print(f"Error in process_stream endpoint: {e}")
        return jsonify({'error': str(e)}), 500

def find_relevant_document_context(field: str, value: str, document_lines: list) -> str:
    """Find relevant context for a field-value pair in document text"""
    import re
    
    if not document_lines:
        return ""
    
    full_text = '\n'.join(document_lines)
    
    # Create search terms
    search_terms = []
    if field:
        field_clean = re.sub(r'[^\w\s]', ' ', field.lower()).strip()
        field_words = [w for w in field_clean.split() if len(w) > 2]
        search_terms.extend(field_words)
    
    if value:
        value_clean = str(value).lower()
        search_terms.append(value_clean)
        # Extract numbers and currencies
        numbers = re.findall(r'\d+\.?\d*', value_clean)
        search_terms.extend(numbers)
    
    # Find sentences containing search terms
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z])', full_text)
    relevant_sentences = []
    
    for sentence in sentences:
        sentence_lower = sentence.lower()
        score = 0
        
        for term in search_terms:
            if term and term in sentence_lower:
                score += 1
        
        if score >= 1 and len(sentence.strip()) > 20:
            relevant_sentences.append(sentence.strip())
    
    # Return top 2 most relevant sentences
    return '. '.join(relevant_sentences[:2])

@app.route('/test_prompt', methods=['POST'])
def test_prompt():
    """Test custom context extraction prompt"""
    try:
        data = request.json
        if not data or 'prompt' not in data:
            return jsonify({'error': 'No prompt provided'}), 400
        
        custom_prompt = data['prompt']
        
        # Get stored file data
        file_data_path = session.get('file_data_path')
        if not file_data_path or not os.path.exists(file_data_path):
            return jsonify({'error': 'No file data found. Please upload and extract a PDF first.'}), 400
        
        # Load file data
        with open(file_data_path, 'rb') as f:
            structured_data = pickle.load(f)
        
        # Set custom prompt
        print(f"🎯 Setting custom prompt (length: {len(custom_prompt)})")
        print(f"Custom prompt preview: {custom_prompt[:200]}...")
        set_custom_context_prompt(custom_prompt)
        
        # Verify prompt was set
        from structured_llm_processor import get_custom_context_prompt
        current_prompt = get_custom_context_prompt()
        print(f"✅ Custom prompt verification: {current_prompt is not None}")
        if current_prompt:
            print(f"Stored prompt matches: {current_prompt == custom_prompt}")
        
        # Process with custom prompt
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            print(f"🚀 Starting async processing with custom prompt: {get_custom_context_prompt() is not None}")
            result = loop.run_until_complete(
                process_structured_data_with_llm_async(structured_data)
            )
            print(f"🏁 Async processing completed")
        finally:
            loop.close()
        
        # Extract context data for display BEFORE clearing the prompt
        context_data = []
        performance_stats = {}
        
        if 'enhanced_data_with_comprehensive_context' in result:
            enhanced_data = result['enhanced_data_with_comprehensive_context']
            
            # Extract performance statistics
            performance_stats = get_optimization_stats(enhanced_data)
            
            for entity in enhanced_data:
                context_data.append({
                    'field': entity.get('field', 'Unknown'),
                    'value': entity.get('value', ''),
                    'context': entity.get('context', entity.get('Context', ''))
                })
        
        print(f"🔍 BEFORE CLEARING: Custom prompt still set: {get_custom_context_prompt() is not None}")
        print(f"🔍 Extracted {len(context_data)} entities with contexts")
        
        # Clear the custom prompt to restore default AFTER processing
        clear_custom_context_prompt()
        
        print(f"🔍 AFTER CLEARING: Custom prompt cleared: {get_custom_context_prompt() is None}")
        
        return jsonify({
            'success': True,
            'context_data': context_data,
            'total_entities': len(context_data),
            'performance_stats': performance_stats
        })
        
    except Exception as e:
        # Make sure to clear custom prompt on error
        try:
            clear_custom_context_prompt()
        except:
            pass  # Ignore if function not available
        print(f"Error testing prompt: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/export_xlsx', methods=['POST'])
def export_xlsx():
    """Export CSV data to XLSX format with specific sheet name"""
    try:
        print("🔄 XLSX export endpoint called")
        
        import csv
        from io import StringIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.json
        print(f"📊 Request data received: {bool(data)}")
        
        if not data:
            print("❌ No data provided in request")
            return jsonify({'error': 'No data provided'}), 400
            
        csv_data = data.get('csv_data', '')
        print(f"📄 CSV data length: {len(csv_data) if csv_data else 0}")
        
        if not csv_data:
            print("❌ No CSV data in request")
            return jsonify({'error': 'No CSV data provided'}), 400
        
        print("✅ Starting XLSX generation...")
        
        # Create workbook with specific sheet name
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Extracted_data_comments"
        
        # Parse CSV data properly handling multi-line content and long contexts
        if ws is not None:
            # Use proper CSV parsing to handle multi-line quoted fields
            csv_reader = csv.reader(StringIO(csv_data))
            
            for row_idx, row_line in enumerate(csv_reader, 1):
                # Check if this is a valid row (starts with "rowN:")
                if row_line and len(row_line) > 0 and row_line[0].startswith('row'):
                    # Extract actual data after "rowN: "
                    first_cell = row_line[0]
                    colon_idx = first_cell.find(': ')
                    if colon_idx != -1:
                        # Remove "rowN: " prefix from first column
                        row_line[0] = first_cell[colon_idx + 2:]
                        
                        # Add to worksheet
                        for col_idx, cell_value in enumerate(row_line, 1):
                            # Handle multi-line content in Excel cells
                            if isinstance(cell_value, str):
                                if '\\n' in cell_value:
                                    cell_value = cell_value.replace('\\n', '\n')
                                # Ensure no truncation of long contexts
                                if len(cell_value) > 32767:  # Excel cell limit
                                    # Split long content but keep it readable
                                    cell_value = cell_value[:32760] + "..."
                            
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                            
                            # Enable text wrapping for multi-line content, especially Context column
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            # Special handling for Context column (usually last column)
                            if col_idx == len(row_line) and row_idx > 1:  # Context column, not header
                                # Set larger row height for context
                                ws.row_dimensions[row_idx].height = max(60, len(str(cell_value)) // 100 * 15)
                            
                            # Style header row
                            if row_idx == 1:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
            # Auto-adjust column widths with special handling for Context column
            try:
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = column[0].column_letter
                    is_context_column = False
                    
                    # Check if this is the Context column (usually the last column)
                    if col_idx == ws.max_column:
                        header_cell = ws.cell(row=1, column=col_idx)
                        if header_cell.value and 'context' in str(header_cell.value).lower():
                            is_context_column = True
                    
                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    if is_context_column:
                        # Context column gets wider width to accommodate longer text
                        adjusted_width = min(max_length // 4 + 20, 100)  # Wider for context, cap at 100
                    else:
                        # Regular columns
                        adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters for regular columns
                    
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as width_error:
                print(f"Error adjusting column widths: {width_error}")
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        print("✅ XLSX file generated successfully, sending to client")
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='extracted_data_comments.xlsx'
        )
        
    except Exception as e:
        print(f"❌ XLSX export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to generate XLSX: {str(e)}'}), 500

@app.route('/export_csv', methods=['POST'])
def export_csv():
    """Export as CSV file"""
    try:
        data = request.json
        if not data or 'csv_data' not in data:
            return jsonify({'error': 'No CSV data provided'}), 400
        
        csv_data = data['csv_data']
        
        return send_file(
            io.BytesIO(csv_data.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='extracted_data_comments.csv'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate CSV: {str(e)}'}), 500

if __name__ == '__main__':
    # Production configuration
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') == 'development'
    
    app.run(
        host='0.0.0.0',
        port=port,
        debug=debug,
        use_reloader=False  # Disable reloader for production
    )