import os
import shutil
import tempfile
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import pickle

# Import processing modules
from textract_processor import extract_structured_data_from_pdf_bytes
from structured_llm_processor_simple import process_structured_data_with_llm_async
import asyncio

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Store processed data temporarily
session_data = {}

@app.route('/')
def index():
    return render_template('index_simple.html')

@app.route('/extract', methods=['POST'])
def extract():
    """Extract structured data from uploaded PDF"""
    try:
        print("Extract endpoint called")
        print(f"Request files: {list(request.files.keys())}")
        
        if 'pdf' not in request.files:
            return jsonify({'error': 'No PDF file provided'}), 400
        
        file = request.files['pdf']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        print(f"File received: {file.filename}, size: {len(file.read())}")
        file.seek(0)  # Reset file pointer
        
        # Read PDF bytes
        pdf_bytes = file.read()
        print(f"PDF bytes read: {len(pdf_bytes)} bytes")
        
        # Extract structured data
        print("Calling extract_structured_data_from_pdf_bytes...")
        structured_data = extract_structured_data_from_pdf_bytes(pdf_bytes)
        print(f"Structured data extracted: {len(structured_data)} items")
        
        # Store in session (using simple in-memory storage)
        session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_data[session_id] = structured_data
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'data': structured_data,
            'message': f'Successfully extracted {len(structured_data)} data items'
        })
        
    except Exception as e:
        print(f"Error in extract endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/process', methods=['POST'])
def process():
    """Process structured data with AI to add context"""
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        print("Starting AI processing...")
        
        # Process with LLM asynchronously
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            result = loop.run_until_complete(
                process_structured_data_with_llm_async(data)
            )
        finally:
            loop.close()
        
        print("AI processing completed")
        
        # Convert to CSV format for export
        csv_data = convert_to_csv(result)
        
        return jsonify({
            'success': True,
            'result': result,
            'csv_data': csv_data,
            'message': 'Processing completed successfully'
        })
        
    except Exception as e:
        print(f"Error in process endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def convert_to_csv(result):
    """Convert processed result to CSV format"""
    csv_rows = ['Field,Value,Context']
    
    if 'enhanced_data_with_comprehensive_context' in result:
        for entity in result['enhanced_data_with_comprehensive_context']:
            field = entity.get('field', 'Unknown')
            value = entity.get('value', '')
            context = entity.get('context', entity.get('Context', ''))
            
            # Clean and escape CSV values
            field = str(field).replace('"', '""')
            value = str(value).replace('"', '""')
            context = str(context).replace('"', '""')
            
            csv_rows.append(f'"{field}","{value}","{context}"')
    
    return '\n'.join(csv_rows)

@app.route('/export_xlsx', methods=['POST'])
def export_xlsx():
    """Export CSV data to XLSX format"""
    try:
        print("🔄 XLSX export endpoint called")
        
        import io
        import csv
        from io import StringIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.json
        if not data or 'csv_data' not in data:
            return jsonify({'error': 'No CSV data provided'}), 400
        
        csv_data = data['csv_data']
        print(f"📄 Processing CSV data, length: {len(csv_data)}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted_Data_Context"
        
        # Parse CSV and add to worksheet
        csv_reader = csv.reader(StringIO(csv_data))
        
        for row_idx, row in enumerate(csv_reader, 1):
            for col_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Style header row
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print("✅ XLSX file generated successfully")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='extracted_data_with_context.xlsx'
        )
        
    except Exception as e:
        print(f"❌ XLSX export error: {e}")
        return jsonify({'error': f'Failed to generate XLSX: {str(e)}'}), 500

@app.route('/export_csv', methods=['POST'])
def export_csv():
    """Export as CSV file"""
    try:
        data = request.json
        if not data or 'csv_data' not in data:
            return jsonify({'error': 'No CSV data provided'}), 400
        
        csv_data = data['csv_data']
        
        # Create temporary file
        output = io.StringIO()
        output.write(csv_data)
        output.seek(0)
        
        return send_file(
            io.BytesIO(csv_data.encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name='extracted_data_with_context.csv'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate CSV: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)